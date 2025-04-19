
import os
st.write("Tesseract path:", pytesseract.pytesseract.tesseract_cmd)

import streamlit as st
from PIL import Image, ImageOps, ImageEnhance
import pytesseract
import fitz
import io
import pandas as pd
import os
import json
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", "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"))

st.set_page_config(page_title="QuantumDocs: Sorted OCR", layout="wide")
st.title("üî∑ QuantumDocs: Sorted OCR by Filename")

uploaded_files = st.file_uploader("üìÅ Upload documents (processed in filename order)", type=["pdf", "png", "jpg", "jpeg", "tif"], accept_multiple_files=True)

def is_dark_image(img):
    grayscale = img.convert("L")
    hist = grayscale.histogram()
    brightness = sum(i * hist[i] for i in range(256)) / (img.width * img.height)
    return brightness < 128

def prepare_image(img):
    if is_dark_image(img):
        inverted = ImageOps.invert(img.convert("RGB"))
        enhanced = ImageEnhance.Contrast(inverted).enhance(1.5)
        return enhanced
    return img.convert("RGB")

def extract_text(file):
    if file.type == "application/pdf":
        text = ""
        doc = fitz.open(stream=file.read(), filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            cleaned = prepare_image(img)
            text += pytesseract.image_to_string(cleaned) + "\n"
        return text
    else:
        img = Image.open(file)
        cleaned = prepare_image(img)
        return pytesseract.image_to_string(cleaned)

def extract_fields(ocr_text):
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a legal title assistant. Return fields in JSON."},
                {"role": "user", "content": f"""Extract the following from this oil and gas title document and return as JSON:
- Document Type
- Grantor
- Grantee
- Date of Instrument
- Legal Description
- Clauses

For the 'Clauses' field, include any sentences that contain clause-indicative language such as:
- reserving
- hereby reserve
- less and except
- save and except
- reserves unto
- excepts and reserves
- subject to
- burdened by
- deducted from
- excluding all oil, gas, and other minerals
- excepting oil, gas, and other hydrocarbons
- any clause that includes the phrase: an undivided [fraction or percentage] of its right, title and interest

Document text:
{ocr_text}"""}
            ]
        )
        raw = response.choices[0].message.content.strip()
        if "{" in raw and "}" in raw:
            json_block = raw[raw.index("{"):raw.rindex("}")+1]
            try:
                fields = json.loads(json_block)
                for k, v in fields.items():
                    if isinstance(v, list):
                        fields[k] = ", ".join(str(x) for x in v)
                return fields
            except Exception as e:
                st.warning(f"JSON parse error: {e}")
                st.text_area("Raw GPT Output", raw, height=300)
                return {}
        else:
            st.warning("No valid JSON structure found.")
            st.text_area("Raw GPT Output", raw, height=300)
            return {}
    except Exception as e:
        st.error(f"GPT request error: {e}")
        return {}

if uploaded_files:
    uploaded_files = sorted(uploaded_files, key=lambda x: x.name.lower())
    progress = st.progress(0)
    results = []

    for i, file in enumerate(uploaded_files):
        st.header(f"üìÑ {file.name}")
        file_key = f"doc_{file.name}"

        if file_key not in st.session_state:
            ocr_text = extract_text(file)
            fields = extract_fields(ocr_text)
            st.session_state[file_key] = {"ocr": ocr_text, "fields": fields}
        else:
            ocr_text = st.session_state[file_key]["ocr"]
            fields = st.session_state[file_key]["fields"]

        doc_type = st.text_input("Document Type", value=fields.get("Document Type", ""), key=f"doc_type_{file.name}")
        grantor = st.text_input("Grantor", value=fields.get("Grantor", ""), key=f"grantor_{file.name}")
        grantee = st.text_input("Grantee", value=fields.get("Grantee", ""), key=f"grantee_{file.name}")
        date_instr = st.text_input("Date of Instrument", value=fields.get("Date of Instrument", ""), key=f"date_{file.name}")
        legal_desc = st.text_area("Legal Description", value=fields.get("Legal Description", ""), key=f"legal_{file.name}", height=80)
        clauses = st.text_area("Clauses", value=fields.get("Clauses", ""), key=f"clauses_{file.name}", height=80)

        st.subheader("üìú Extracted OCR Text")
        st.text_area("OCR Text", ocr_text, height=200, key=f"ocr_{file.name}")

        st.subheader("ü§ñ Ask QuantumDocs")
        user_q = st.text_input("Ask a question:", key=f"q_{file.name}")
        if st.button(f"Submit - {file.name}"):
            try:
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are a legal title assistant specializing in oil and gas documents."},
                        {"role": "user", "content": f"Document text:\n{ocr_text}\n\nQuestion: {user_q}"}
                    ]
                )
                st.success(response.choices[0].message.content)
            except Exception as e:
                st.error(f"Error: {e}")

        results.append({
            "Filename": file.name,
            "Document Type": doc_type,
            "Grantor": grantor,
            "Grantee": grantee,
            "Date of Instrument": date_instr,
            "Legal Description": legal_desc,
            "Clauses": clauses
        })

        progress.progress((i + 1) / len(uploaded_files))

    df = pd.DataFrame(results)
    xlsx_path = "quantumdocs_sorted_output.xlsx"
    df.to_excel(xlsx_path, index=False)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(15, min(80, max_length + 2))
    wb.save(xlsx_path)

    st.download_button("üì• Download Excel Output", data=open(xlsx_path, "rb"), file_name="quantumdocs_sorted_output.xlsx")

else:
    st.info("Upload document(s) to begin.")
